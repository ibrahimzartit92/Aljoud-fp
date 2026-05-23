import io
import os
import arabic_reshaper
from bidi.algorithm import get_display
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont


ARABIC_FONT_NAME = "NotoNaskhArabic"
ARABIC_FONT_REL_PATH = "server/app/static/fonts/NotoNaskhArabic-Regular.ttf"
ARABIC_FONT_PATH = os.path.abspath(
    os.path.join(os.path.dirname(__file__), "..", "static", "fonts", "NotoNaskhArabic-Regular.ttf")
)
_ARABIC_FONT_REGISTERED = False


def _ensure_arabic_pdf_font():
    global _ARABIC_FONT_REGISTERED
    if _ARABIC_FONT_REGISTERED:
        return
    if not os.path.exists(ARABIC_FONT_PATH):
        raise RuntimeError(f"Missing Arabic PDF font: {ARABIC_FONT_REL_PATH}")
    pdfmetrics.registerFont(TTFont(ARABIC_FONT_NAME, ARABIC_FONT_PATH))
    _ARABIC_FONT_REGISTERED = True


def pdf_ar(text) -> str:
    s = "" if text is None else str(text)
    if not s:
        return ""
    return get_display(arabic_reshaper.reshape(s))


def _autosize_worksheet(ws, max_width=55):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            v = "" if cell.value is None else str(cell.value)
            if len(v) > max_len:
                max_len = len(v)
        ws.column_dimensions[col_letter].width = min(max_width, max(10, max_len + 2))


def export_hours_excel(rows_or_pack, header_text: str):
    """
    Supports:
      - list[dict] (legacy)
      - dict pack: {"sessions": [...], "daily":[...], "totals":[...], "grand": {...}}
    Produces 3 sheets for readability.
    """
    wb = Workbook()

    def write_sheet(title: str, rows: list[dict]):
        ws = wb.create_sheet(title=title)
        ws.sheet_view.rightToLeft = True
        ws.append([header_text or "تقرير ساعات العمل"])
        ws.append([])
        if rows:
            keys = list(rows[0].keys())
            ws.append(keys)
            for r in rows:
                ws.append([r.get(k, "") for k in keys])
        _autosize_worksheet(ws)

    # remove default sheet
    wb.remove(wb.active)

    if isinstance(rows_or_pack, dict):
        sessions = rows_or_pack.get("sessions") or []
        daily = rows_or_pack.get("daily") or []
        totals = rows_or_pack.get("totals") or []
        grand = rows_or_pack.get("grand") or {}

        write_sheet("اليومي", daily)
        write_sheet("الإجماليات", totals)
        write_sheet("الجلسات", sessions)

        # small grand row on Totals sheet
        if totals:
            ws = wb["الإجماليات"]
            ws.append([])
            ws.append(["الإجمالي العام", "", grand.get("net_time_txt", "00:00")])
            _autosize_worksheet(ws)

    else:
        # legacy list
        write_sheet("الجلسات", rows_or_pack or [])

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def export_hours_pdf(rows_or_pack, header_text: str):
    """
    Supports list[dict] or dict pack.
    Uses ReportLab tables (readable) and DOES NOT add an extra blank page.
    """
    _ensure_arabic_pdf_font()
    out = io.BytesIO()
    doc = SimpleDocTemplate(out, pagesize=landscape(A4), leftMargin=20, rightMargin=20, topMargin=20, bottomMargin=20)
    styles = getSampleStyleSheet()
    for style_name in ("Title", "Heading2", "Heading3", "Normal"):
        styles[style_name].fontName = ARABIC_FONT_NAME
    story = []

    story.append(Paragraph(pdf_ar(header_text or "تقرير ساعات العمل"), styles["Title"]))
    story.append(Spacer(1, 10))

    def add_table(title: str, rows: list[dict], max_rows=45):
        story.append(Paragraph(pdf_ar(title), styles["Heading2"]))
        story.append(Spacer(1, 6))
        if not rows:
            story.append(Paragraph(pdf_ar("لا توجد بيانات"), styles["Normal"]))
            story.append(Spacer(1, 10))
            return

        keys = list(rows[0].keys())
        data = [[pdf_ar(k) for k in keys]]
        for r in rows[:max_rows]:
            data.append([pdf_ar(r.get(k, "")) for k in keys])

        tbl = Table(data, repeatRows=1)
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#222222")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, -1), ARABIC_FONT_NAME),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.lightgrey]),
        ]))
        story.append(tbl)
        if len(rows) > max_rows:
            story.append(Spacer(1, 6))
            story.append(Paragraph(pdf_ar(f"يتم عرض أول {max_rows} صف من أصل {len(rows)}"), styles["Normal"]))
        story.append(Spacer(1, 14))

    if isinstance(rows_or_pack, dict):
        add_table("الإجماليات اليومية لكل موظف", rows_or_pack.get("daily") or [])
        add_table("إجماليات الموظفين", rows_or_pack.get("totals") or [])

        grand = rows_or_pack.get("grand") or {}
        story.append(Paragraph(
            pdf_ar(f"الإجمالي العام (صافي الوقت): {grand.get('net_time_txt','00:00')}"),
            styles["Heading3"],
        ))
        story.append(Spacer(1, 14))

        story.append(PageBreak())
        add_table("تفاصيل الجلسات", rows_or_pack.get("sessions") or [], max_rows=60)
    else:
        add_table("الجلسات", rows_or_pack or [], max_rows=60)

    doc.build(story)
    out.seek(0)
    return out
